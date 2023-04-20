from openpyxl import load_workbook

# Loading excel file and sheets
wb = load_workbook(filename='hop_list.xlsx')
ws = wb.active
nz = wb['New Zealand']
usa = wb['American']
au = wb['Australia']
pl = wb['Polska']


def get_hop_name():
    """Function that asks user for hop name and returns it"""
    hop_name = input('Wprowadz nazwe chmielu: ')
    hop_name = hop_name.strip()
    return hop_name


def find_hop_loop(sheet, hop_name):
    """Function that finds wanted hop in main excel sheet, navigates to appropriate country sheet"""
    val = False
    country = sheet.cell(row=1, column=1)
    for row in sheet:
        for cell in row:
            if cell.value == hop_name.title():
                country = sheet.cell(row=cell.row, column=cell.column + 1)
                val = True
            # Special cases
            elif (hop_name.lower() == 'wai-iti' or hop_name.lower() == 'waiiti' or hop_name.lower() == 'wai iti')\
                    and cell.value == 'Wai-iti':
                country = sheet.cell(row=cell.row, column=cell.column + 1)
                val = True
            elif (hop_name.lower() == 'columbus' or hop_name.lower() == 'tomahawk' or
                  hop_name.lower() == 'zeus' or hop_name.lower() == 'ctz') and cell.value == 'Columbus/Tomahawk/Zeus':
                country = sheet.cell(row=cell.row, column=cell.column + 1)
                val = True
            elif hop_name.lower() == 'pulawski' and cell.value == 'Puławski':
                country = sheet.cell(row=cell.row, column=cell.column + 1)
                val = True
            elif hop_name.lower() == 'idaho' and cell.value == 'Idaho 7':
                country = sheet.cell(row=cell.row, column=cell.column + 1)
                val = True
            elif (hop_name.lower() == 'nelson' or hop_name.lower() == 'sauvin') and cell.value == 'Nelson Sauvin':
                country = sheet.cell(row=cell.row, column=cell.column + 1)
                val = True
    return val, country


def get_sheet_name(country):
    """Function that returns appropriate sheet name based on country name"""
    if country == 'Nowa Zelandia':
        return nz
    elif country == 'USA':
        return usa
    elif country == 'Australia':
        return au
    elif country == 'Polska':
        return pl
    else:
        print('Nie ma takiego kraju w bazie!')


def find_hop(sheet):
    """Function that finds wanted hop in main excel sheet, returns wanted sheet and hop name"""
    hop_name = get_hop_name()
    val, country = find_hop_loop(sheet, hop_name)
    if not val:
        print('Nie ma takiego chmielu w bazie!')
    else:
        hop_sheet = get_sheet_name(country.value)
        return hop_sheet, hop_name


def print_aromas(sheet, cell):
    gen_aromas = []  # list of general aromas
    spec_aromas = []  # list of specific aromas
    count = 0  # counter to know how many rows to go back to the initial cell.row number
    print(f'\n{sheet.cell(row=cell.row, column=cell.column).value}\n')
    print('Aromaty ogólne: ')
    # Read from cell below in Aromaty ogólnie column until 'Typ chmielu' is found
    while True:
        if sheet.cell(row=cell.row, column=cell.column + 1).value == 'Typ chmielu':
            break
        val1 = sheet.cell(row=cell.row, column=cell.column + 1).value
        gen_aromas.append(val1)
        print(f'\t- {val1}')
        count += 1
        cell.row += 1

    cell.row -= count
    count = 0
    print("-" * 30)
    print("Aromaty 'dokladnie': ")
    # Read from cell below in Aromaty 'dokladnie' column until None is found
    while True:
        if sheet.cell(row=cell.row, column=cell.column + 2).value is None:
            break
        val2 = sheet.cell(row=cell.row, column=cell.column + 2).value
        print(f'\t- {val2}')
        spec_aromas.append(val2)
        cell.row += 1
        count += 1
    cell.row -= count
    return gen_aromas, spec_aromas


def get_aromas(sheet, hop_name):
    """Function that returns aromas of wanted hop"""
    for row in sheet:
        for cell in row:
            if cell.value == hop_name.title():
                print_aromas(sheet, cell)
                break
            # Special cases
            elif (hop_name.lower() == 'wai-iti' or hop_name.lower() == 'waiiti' or hop_name.lower() == 'wai iti') \
                    and cell.value == 'Wai-iti':
                print_aromas(sheet, cell)
                break
            elif (hop_name.lower() == 'columbus' or hop_name.lower() == 'tomahawk' or
                  hop_name.lower() == 'zeus' or hop_name.lower() == 'ctz') and cell.value == 'CTZ':
                print_aromas(sheet, cell)
                break
            elif hop_name.lower() == 'pulawski' and cell.value == 'Puławski':
                print_aromas(sheet, cell)
                break
            elif hop_name.lower() == 'idaho' and cell.value == 'Idaho 7':
                print_aromas(sheet, cell)
                break
            elif (hop_name.lower() == 'nelson' or hop_name.lower() == 'sauvin') and cell.value == 'Nelson Sauvin':
                print_aromas(sheet, cell)
                break


def main():
    """Main function"""
    while True:
        data = find_hop(ws)  # data[0] - sheet name / data[1] - hop name
        if data:
            get_aromas(data[0], data[1])

        answer = input('\nCzy chcesz wyszukac kolejny chmiel? (t/n) ')
        if answer == 'n':
            break


if __name__ == '__main__':
    main()

